"""
Excel export module for transaction data
"""
import io
import logging
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export transaction data to Excel files"""

    # Define columns in order
    COLUMNS = ['date', 'description', 'debit amt', 'credit amt', 'balance']

    # Column widths
    COLUMN_WIDTHS = {
        'date': 12,
        'description': 35,
        'debit amt': 12,
        'credit amt': 12,
        'balance': 12
    }

    def __init__(self):
        if not Workbook:
            raise ImportError("openpyxl is required for Excel export")

    def create_workbook(self, transactions: List[Dict[str, Any]], bank: str = None) -> bytes:
        """
        Create Excel workbook from transaction data
        Args:
            transactions: List of transaction dicts with keys: date, description, debit, credit, balance
            bank: Optional bank name for title
        Returns:
            Excel file as bytes
        """
        logger.info(f"Creating Excel workbook with {len(transactions)} transactions")

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Add title
        if bank:
            ws['A1'] = f"{bank} Bank Statement"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:E1')
            row_offset = 3
        else:
            row_offset = 1

        # Add headers
        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.value = col_name.upper()
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_idx, transaction in enumerate(transactions, row_offset + 1):
            for col_idx, col_name in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Get value from transaction dict (handle both snake_case and full names)
                value = transaction.get(col_name)
                if value is None:
                    # Try alternative names
                    if col_name == 'debit amt':
                        value = transaction.get('debit')
                    elif col_name == 'credit amt':
                        value = transaction.get('credit')

                cell.value = value

                # Format cells
                if col_name == 'date':
                    cell.alignment = Alignment(horizontal="center")
                elif col_name in ['debit amt', 'credit amt', 'balance']:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # Set column widths
        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = self.COLUMN_WIDTHS.get(col_name, 15)

        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=row_offset, max_row=row_offset + len(transactions),
                                min_col=1, max_col=len(self.COLUMNS)):
            for cell in row:
                cell.border = thin_border

        # Save to bytes
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)

        logger.info(f"Excel workbook created successfully: {len(transactions)} rows")
        return bytes_io.getvalue()

    def save_to_file(self, transactions: List[Dict[str, Any]], output_path: str, bank: str = None):
        """
        Save transactions to Excel file
        Args:
            transactions: List of transaction dicts
            output_path: Path where to save the file
            bank: Optional bank name for title
        """
        excel_bytes = self.create_workbook(transactions, bank)

        with open(output_path, 'wb') as f:
            f.write(excel_bytes)

        logger.info(f"Excel file saved: {output_path}")
        return output_path


def export_to_excel(transactions: List[Dict[str, Any]], output_path: str = None, bank: str = None) -> bytes:
    """
    Export transaction list to Excel
    Args:
        transactions: List of transaction dicts
        output_path: Optional path to save file (if None, returns bytes)
        bank: Optional bank name
    Returns:
        Excel file as bytes
    """
    exporter = ExcelExporter()

    if output_path:
        exporter.save_to_file(transactions, output_path, bank)
        with open(output_path, 'rb') as f:
            return f.read()
    else:
        return exporter.create_workbook(transactions, bank)
